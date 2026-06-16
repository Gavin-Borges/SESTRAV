"""
SESTRAV IEDB Data Loading and Cleaning Module

Handles TWO IEDB export formats:
  1. Epitope Table exports — peptides in column 3 ("Name" sub-header),
     immunogenicity labels derived from FILENAME ("T-cell positive" vs
     "T-cell negative"), NO allele column, NO "Qualitative Measure" column.
     The HPV16 negative file ("HVP16" typo) has 16 columns and no sub-header
     row; all other Epitope Table files have 32 columns with a multi-row header.
  2. T-cell Assay exports (legacy support) — "Description" column for peptides,
     "Qualitative Measure" column for labels, optional "Allele" column.

Downstream processing: 8-11mer filtering, standard amino acid validation,
per-peptide duplicate resolution by majority vote, gold-standard hold-out.
"""

import os
import sys
import pandas as pd
import openpyxl

STANDARD_AA = set('ACDEFGHIKLMNPQRSTVWY')

GOLD_STANDARD_EPITOPES = [
    'CLGGLLTMV', 'GLCTLVAML', 'FLRGRAYGL', 'FLRGRAYGI', 'RAKFKQLL',
    'IVTDFSVIK', 'RPPIFIRRL', 'HPVGEADYFEY', 'TYSAGIVQI', 'AVFDRKSDAK',
    'YVLDHLIVV', 'YMLDLQPET', 'RAHYNIVTF', 'LLMGTLGIV', 'KLPQLCTEL',
    'TIHDIILECV'
]
# Note: this holdout list has 16 entries because both FLR strain variants
# (FLRGRAYGI and FLRGRAYGL) are excluded from training.

MHC_CLASS_II_PREFIXES = ('HLA-DR', 'HLA-DP', 'HLA-DQ')


def _label_from_filename(filename):
    """Derive binary immunogenicity label from IEDB Epitope Table filename.

    IEDB Epitope Table exports encode the T-cell assay outcome in the
    filename only (e.g. "T-cell positive" or "T-cell negative").  There
    is no per-row label column inside these files.
    """
    fname_lower = filename.lower()
    if 'positive' in fname_lower:
        return 1
    if 'negative' in fname_lower:
        return 0
    return None


def _virus_from_filename(filename):
    """Determine virus type from filename, handling the HVP16 typo."""
    fname_lower = filename.lower()
    if 'ebv' in fname_lower:
        return 'EBV'
    if 'hpv16' in fname_lower or 'hvp16' in fname_lower:
        return 'HPV16'
    if 'hpv11' in fname_lower:
        return 'HPV11'
    return None


def _detect_format(filepath):
    """Detect whether an xlsx file is Epitope Table or T-cell Assay format.

    Returns ('epitope_table', has_subheader) or ('tcell_assay', None).
    """
    if filepath.endswith('.csv'):
        df = pd.read_csv(filepath, nrows=3, header=None)
        for _, row in df.iterrows():
            if any('qualitative' in str(c).lower() for c in row):
                return 'tcell_assay', None
        return 'epitope_table', False

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    max_col = ws.max_column or 50
    row1 = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    row2 = [ws.cell(row=2, column=c).value for c in range(1, max_col + 1)]
    wb.close()

    row1_lower = [str(v).lower() if v else '' for v in row1]
    row2_lower = [str(v).lower() if v else '' for v in row2]

    if any('qualitative' in c for c in row1_lower):
        return 'tcell_assay', None
    if any('qualitative' in c for c in row2_lower):
        return 'tcell_assay', None

    has_subheader = (
        row2[0] is not None
        and isinstance(row2[0], str)
        and row2[0].strip().upper().startswith('IEDB')
    )
    return 'epitope_table', has_subheader


def _load_epitope_table(filepath, has_subheader):
    """Load an IEDB Epitope Table export, handling multi-row headers.

    Peptide sequences are always in the 3rd column (0-indexed col 2).
    Files with a sub-header row have data starting at row 3 (0-indexed);
    files without (HPV16 negative) have data starting at row 2.

    Returns a list of dicts with keys: peptide, and optionally
    antigen_name, organism_name extracted from Epitope Table columns.
    """
    if filepath.endswith('.csv'):
        skip = 2 if has_subheader else 1
        df = pd.read_csv(filepath, skiprows=skip, header=None)
    else:
        skip = 2 if has_subheader else 1
        df = pd.read_excel(filepath, skiprows=skip, header=None)

    peptide_col_idx = 2
    # IEDB Epitope Table 32-column layout (sub-header files):
    #   col 2 = "Name" (peptide sequence)
    #   col 6 = "Antigen Name" (source protein, e.g. "Protein E7")
    #   col 8 = "Organism Name" (e.g. "Human papillomavirus type 16")
    # 16-column files (no sub-header) use the same col-2 for peptides
    # but may have antigen/organism at different offsets.
    antigen_col_idx = 6 if df.shape[1] > 6 else None
    organism_col_idx = 8 if df.shape[1] > 8 else None

    records = []
    for _, row in df.iterrows():
        val = row.iloc[peptide_col_idx] if len(row) > peptide_col_idx else None
        if val is not None and isinstance(val, str):
            seq = val.strip().upper()
        else:
            seq = None

        antigen = None
        if antigen_col_idx is not None and len(row) > antigen_col_idx:
            raw = row.iloc[antigen_col_idx]
            if raw is not None and isinstance(raw, str) and raw.strip():
                antigen = raw.strip()

        organism = None
        if organism_col_idx is not None and len(row) > organism_col_idx:
            raw = row.iloc[organism_col_idx]
            if raw is not None and isinstance(raw, str) and raw.strip():
                organism = raw.strip()

        records.append({
            'peptide': seq,
            'antigen_name': antigen,
            'organism_name': organism,
        })
    return records


def load_iedb_file(filepath):
    """Load a single IEDB file (.xlsx or .csv) into a DataFrame.

    Prevents false-positive multi-header detection when the first data row
    contains the word 'qualitative' (e.g. 'qualitative binding').
    """
    if filepath.endswith('.csv'):
        df_test = pd.read_csv(filepath, nrows=2, header=None)
        row0 = df_test.iloc[0].astype(str).tolist()
        if any(' - ' in c for c in row0):
            return pd.read_csv(filepath, low_memory=False)
        if len(df_test) > 1 and any('qualitative' in str(c).lower() for c in df_test.iloc[1]):
            return pd.read_csv(filepath, header=1, low_memory=False)
        return pd.read_csv(filepath, low_memory=False)
    
    df_test = pd.read_excel(filepath, nrows=2, header=None)
    row0 = df_test.iloc[0].astype(str).tolist()
    if any(' - ' in c for c in row0):
        return pd.read_excel(filepath)
    if len(df_test) > 1 and any('qualitative' in str(c).lower() for c in df_test.iloc[1]):
        return pd.read_excel(filepath, header=1)
    return pd.read_excel(filepath)


def standardize_allele(allele_str):
    """Normalize HLA allele formatting (e.g., A*0201 -> HLA-A*02:01)."""
    if pd.isna(allele_str):
        return None
    allele = str(allele_str).strip()
    if not allele.startswith('HLA-'):
        allele = 'HLA-' + allele
    return allele


def map_label(qualitative_measure):
    """Map IEDB 'Qualitative Measure' to binary label (T-cell Assay format)."""
    if pd.isna(qualitative_measure):
        return None
    val = str(qualitative_measure).strip().lower()
    if val.startswith('positive'):
        return 1
    if val == 'negative':
        return 0
    return None


def is_valid_peptide(seq, min_len=8, max_len=11):
    """Check peptide is standard amino acids only and within MHC-I length range."""
    if pd.isna(seq):
        return False
    seq = str(seq).strip().upper()
    if not (min_len <= len(seq) <= max_len):
        return False
    return all(aa in STANDARD_AA for aa in seq)


def is_mhc_class_i(allele_str):
    """Reject MHC class II alleles (HLA-DR, HLA-DP, HLA-DQ)."""
    if pd.isna(allele_str):
        return False
    return not str(allele_str).strip().startswith(MHC_CLASS_II_PREFIXES)


def _infer_protein_gene(antigen_name, virus):
    """Map IEDB antigen names to standardized protein/gene identifiers.

    IEDB Epitope Table "Antigen Name" values are free-text (e.g. "Protein E7",
    "Latent membrane protein 2", "Nuclear antigen EBNA-3A").  This mapper
    normalizes the most common entries for EBV and HPV16 to short canonical
    gene symbols used elsewhere in SESTRAV (e.g. "E7", "LMP2A", "EBNA3A").
    """
    if antigen_name is None:
        return None
    name = antigen_name.strip()
    name_lower = name.lower()

    if virus in ('HPV16', 'HPV11'):
        for gene in ('E1', 'E2', 'E4', 'E5', 'E6', 'E7', 'L1', 'L2'):
            if gene.lower() in name_lower:
                return gene
        return name

    if virus == 'EBV':
        ebv_map = {
            'lmp1': 'LMP1', 'lmp2': 'LMP2A', 'lmp-2': 'LMP2A',
            'lmp 2': 'LMP2A', 'latent membrane protein 1': 'LMP1',
            'latent membrane protein 2': 'LMP2A',
            'ebna1': 'EBNA1', 'ebna-1': 'EBNA1', 'ebna 1': 'EBNA1',
            'ebna2': 'EBNA2', 'ebna-2': 'EBNA2',
            'ebna3a': 'EBNA3A', 'ebna-3a': 'EBNA3A', 'ebna3': 'EBNA3A',
            'ebna3b': 'EBNA3B', 'ebna-3b': 'EBNA3B',
            'ebna3c': 'EBNA3C', 'ebna-3c': 'EBNA3C',
            'bzlf1': 'BZLF1', 'brlf1': 'BRLF1',
            'bmlf1': 'BMLF1', 'sm protein': 'BMLF1',
            'gp350': 'GP350', 'gp340': 'GP350',
            'nuclear antigen': None,
        }
        for pattern, gene in ebv_map.items():
            if pattern in name_lower:
                if gene is not None:
                    return gene
                # "nuclear antigen" needs further disambiguation
                for suffix in ('3a', '3b', '3c', '1', '2', '3'):
                    if suffix in name_lower:
                        return 'EBNA' + suffix.upper()
                return 'EBNA'
        return name

    return name


def _infer_strain(organism_name, virus):
    """Extract strain info from IEDB organism name where possible.

    IEDB "Organism Name" examples:
      "Human herpesvirus 4"                 -> None (generic EBV)
      "Human herpesvirus 4 strain B95-8"    -> "B95-8"
      "Human papillomavirus type 16"        -> "HPV16"
      "Human papillomavirus type 18"        -> "HPV18"
    """
    if organism_name is None:
        return None
    name_lower = organism_name.lower()

    if virus == 'EBV':
        if 'b95-8' in name_lower or 'b95.8' in name_lower or 'b958' in name_lower:
            return 'B95-8'
        if 'gd1' in name_lower:
            return 'GD1'
        if 'ag876' in name_lower:
            return 'AG876'
        if 'akata' in name_lower:
            return 'Akata'
        return None

    if virus in ('HPV16', 'HPV11'):
        if 'type 16' in name_lower or 'hpv16' in name_lower or 'hpv-16' in name_lower:
            return 'HPV16'
        if 'type 18' in name_lower or 'hpv18' in name_lower or 'hpv-18' in name_lower:
            return 'HPV18'
        if 'type 11' in name_lower or 'hpv11' in name_lower:
            return 'HPV11'
        return None

    return None


def load_and_clean_iedb(data_dir, include_hpv11=False):
    """
    Load all IEDB xlsx/csv files from data_dir, auto-detect format, clean,
    and merge into a single training-ready DataFrame.

    Supports both Epitope Table exports (labels from filename) and T-cell
    Assay exports (labels from Qualitative Measure column).

    Args:
        data_dir: path to directory containing IEDB xlsx/csv files
        include_hpv11: whether to include HPV11 data (non-oncogenic type)

    Returns:
        DataFrame with columns: peptide, label, virus, protein, strain
        (protein and strain may be None when source data lacks metadata)
    """
    all_records = []

    for filename in sorted(os.listdir(data_dir)):
        if not filename.endswith(('.xlsx', '.csv')):
            continue

        file_virus = _virus_from_filename(filename)
        if file_virus == 'HPV11' and not include_hpv11:
            continue

        filepath = os.path.join(data_dir, filename)
        fmt, has_subheader = _detect_format(filepath)

        if fmt == 'epitope_table' and file_virus is None:
            print(f"[IEDB Loader] SKIP {filename}: cannot determine virus from filename for Epitope Table")
            continue

        if fmt == 'epitope_table':
            label = _label_from_filename(filename)
            if label is None:
                print(f"[IEDB Loader] SKIP {filename}: cannot determine label from filename")
                continue

            records = _load_epitope_table(filepath, has_subheader)
            n_added = 0
            for rec in records:
                seq = rec['peptide']
                if seq is None:
                    continue
                if not is_valid_peptide(seq):
                    continue
                protein = _infer_protein_gene(rec.get('antigen_name'), file_virus)
                strain = _infer_strain(rec.get('organism_name'), file_virus)
                all_records.append({
                    'peptide': seq,
                    'label': label,
                    'virus': file_virus,
                    'protein': protein,
                    'strain': strain,
                })
                n_added += 1
            print(f"[IEDB Loader] {filename}: Epitope Table format, "
                  f"label={label} (from filename), {n_added} valid 8-11mers")

        elif fmt == 'tcell_assay':
            df = load_iedb_file(filepath)
            peptide_col = None
            label_col = None
            allele_col = None
            antigen_col = None
            organism_col = None
            for col in df.columns:
                cl = str(col).lower().strip()
                if peptide_col is None and (cl == 'description' or cl == 'name' or ('epitope' in cl and ('linear' in cl or 'name' in cl or 'sequence' in cl))):
                    peptide_col = col
                if label_col is None and 'qualitative' in cl:
                    label_col = col
                if allele_col is None and ('allele' in cl or 'mhc present' in cl or 'mhc restriction - name' in cl):
                    allele_col = col
                if antigen_col is None and ('antigen name' in cl or 'source molecule' in cl):
                    antigen_col = col
                if organism_col is None and ('organism' in cl or 'species' in cl):
                    organism_col = col

            if peptide_col is None or label_col is None:
                print(f"[IEDB Loader] SKIP {filename}: T-cell Assay format but missing required columns (found: {list(df.columns[:5])})")
                continue

            n_added = 0
            for _, row in df.iterrows():
                peptide = str(row[peptide_col]).strip().upper() if pd.notna(row[peptide_col]) else None
                row_label = map_label(row[label_col])
                allele = None
                if allele_col and pd.notna(row.get(allele_col)):
                    allele = standardize_allele(row[allele_col])
                    if allele and not is_mhc_class_i(allele):
                        continue

                if peptide is None or row_label is None:
                    continue
                if not is_valid_peptide(peptide):
                    continue

                antigen_raw = None
                if antigen_col and pd.notna(row.get(antigen_col)):
                    antigen_raw = str(row[antigen_col]).strip()
                organism_raw = None
                if organism_col and pd.notna(row.get(organism_col)):
                    organism_raw = str(row[organism_col]).strip()

                row_virus = file_virus
                if row_virus is None:
                    org_lower = organism_raw.lower() if organism_raw else ""
                    if 'human herpesvirus 4' in org_lower or 'epstein barr' in org_lower or 'epstein-barr' in org_lower or 'ebv' in org_lower:
                        row_virus = 'EBV'
                    elif 'human papillomavirus type 16' in org_lower or 'hpv16' in org_lower or 'hpv-16' in org_lower:
                        row_virus = 'HPV16'
                    elif 'human papillomavirus type 11' in org_lower or 'hpv11' in org_lower or 'hpv-11' in org_lower:
                        if not include_hpv11:
                            continue
                        row_virus = 'HPV11'
                    else:
                        continue

                protein = _infer_protein_gene(antigen_raw, row_virus)
                strain = _infer_strain(organism_raw, row_virus)

                record = {
                    'peptide': peptide,
                    'label': row_label,
                    'virus': row_virus,
                    'protein': protein,
                    'strain': strain,
                }
                if allele:
                    record['allele'] = allele
                all_records.append(record)
                n_added += 1
            print(f"[IEDB Loader] {filename}: T-cell Assay format, {n_added} valid records")

    result_df = pd.DataFrame(all_records)

    if result_df.empty:
        print("[IEDB Loader] WARNING: No valid records loaded from any file")
        return result_df

    dupes = result_df.groupby('peptide')['label']
    resolved = dupes.agg(lambda x: int(x.mean() >= 0.5)).reset_index()
    resolved.columns = ['peptide', 'label']

    # Propagate virus, protein, and strain using first non-null per peptide
    meta_cols = ['virus', 'protein', 'strain']
    for col in meta_cols:
        if col in result_df.columns:
            col_map = (
                result_df.dropna(subset=[col])
                .drop_duplicates('peptide')
                [['peptide', col]]
            )
            resolved = resolved.merge(col_map, on='peptide', how='left')
        else:
            resolved[col] = None

    # Propagate allele if present (from T-cell Assay format)
    if 'allele' in result_df.columns:
        allele_map = (
            result_df.dropna(subset=['allele'])
            .drop_duplicates('peptide')
            [['peptide', 'allele']]
        )
        resolved = resolved.merge(allele_map, on='peptide', how='left')

    n_protein = resolved['protein'].notna().sum()
    n_strain = resolved['strain'].notna().sum()
    
    # Calculate comparative statistics
    total_valid_rows = len(result_df)
    unique_peptides = len(resolved)
    allele_coverage_pct = 0.0
    if 'allele' in resolved.columns:
        allele_coverage_pct = (resolved['allele'].notna().sum() / unique_peptides) * 100
    
    conflicts_by_pep = result_df.groupby('peptide')['label'].nunique()
    conflict_count = (conflicts_by_pep > 1).sum()
    conflict_rate_pct = (conflict_count / unique_peptides) * 100
    
    print("\n" + "="*60)
    print("  IEDB DATA LOADING COMPARISON REPORT")
    print("="*60)
    print(f"  Total Valid Rows (Assays) Ingested : {total_valid_rows}")
    print(f"  Unique Peptides Ingested           : {unique_peptides} (vs. 720 baseline)")
    print(f"  HLA Allele Coverage Percentage     : {allele_coverage_pct:.2f}%")
    print(f"  Duplicate Label Conflict Rate      : {conflict_rate_pct:.2f}%")
    print("-"*60)
    print(f"  Positive Class Prevalence          : {resolved['label'].mean():.2%}")
    print(f"  Metadata coverage: protein={n_protein}/{unique_peptides}, "
          f"strain={n_strain}/{unique_peptides}")
    print("="*60 + "\n")
    
    return resolved


def split_gold_standard(df):
    """
    Separate gold-standard epitopes from training data.
    Returns (train_df, gold_standard_df).
    """
    gs_mask = df['peptide'].isin(GOLD_STANDARD_EPITOPES)
    return df[~gs_mask].reset_index(drop=True), df[gs_mask].reset_index(drop=True)


def load_schmidt_2021(filepath, gold_standard_peptides=None):
    """
    Load and clean Schmidt et al. (2021) Cell Reports Medicine dataset.
    Serves as the hard-negative benchmark.
    Isolates (excludes) the gold-standard epitopes.
    """
    if gold_standard_peptides is None:
        gold_standard_peptides = set(GOLD_STANDARD_EPITOPES)
        
    if not os.path.isfile(filepath):
        print(f"[Schmidt Loader] WARNING: File {filepath} not found. Returning empty DataFrame.")
        return pd.DataFrame(columns=['peptide', 'hla', 'label', 'hard_negative_flag'])
    
    # Ingest Schmidt dataset: expected columns are peptide, HLA, and immunogenicity/label.
    if filepath.endswith('.xlsx'):
        df = pd.read_excel(filepath)
    else:
        df = pd.read_csv(filepath)
    
    # Rename columns to standard schema if they differ
    rename_dict = {}
    for col in df.columns:
        cl = str(col).lower().strip()
        if cl in ('peptide', 'epitope', 'sequence'):
            rename_dict[col] = 'peptide'
        elif cl in ('hla', 'allele', 'hla_allele', 'mhc'):
            rename_dict[col] = 'hla'
        elif cl in ('label', 'immunogenicity', 'class', 'active', 'immunogenic'):
            rename_dict[col] = 'label'
            
    df = df.rename(columns=rename_dict)
    
    # Ensure required columns exist
    for col in ['peptide', 'hla', 'label']:
        if col not in df.columns:
            if col == 'label':
                df['label'] = 0
            elif col == 'hla':
                df['hla'] = 'HLA-A*02:01' # default fallback
            else:
                raise ValueError(f"Schmidt dataset missing required column: {col}")
                
    # Clean and validate peptides
    df['peptide'] = df['peptide'].astype(str).str.strip().str.upper()
    df = df[df['peptide'].apply(is_valid_peptide)].copy()
    
    # Standardize alleles
    df['hla'] = df['hla'].apply(standardize_allele)
    
    # Mark hard negatives
    df['hard_negative_flag'] = 1
    
    # Isolate (exclude) gold standard epitopes
    df = df[~df['peptide'].isin(gold_standard_peptides)].reset_index(drop=True)
    
    print(f"[Schmidt Loader] Ingested {len(df)} peptides from Schmidt 2021 dataset (hard-negatives benchmark).")
    return df[['peptide', 'hla', 'label', 'hard_negative_flag']]


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="SESTRAV IEDB Data Loader")
    parser.add_argument("data_dir", help="Directory containing IEDB xlsx/csv files")
    parser.add_argument("--include-hpv11", action="store_true", help="Include HPV11 data")
    args = parser.parse_args()

    if not os.path.isdir(args.data_dir):
        print(f"ERROR: Directory not found: {args.data_dir}")
        sys.exit(1)

    load_and_clean_iedb(args.data_dir, include_hpv11=args.include_hpv11)
